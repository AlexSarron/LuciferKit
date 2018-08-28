from openpyxl import load_workbook
from openpyxl import Workbook
import os
import time

excel = load_workbook(os.path.join(os.getcwd(), '123.xlsx'))#xlsx文件名
sheet = excel.get_sheet_by_name('日海艾拉')#sheet名
print(sheet.max_row)
print(sheet.max_column)
# sheet['a19'] = '1234555'#修改某个cell的值
# value = sheet['a19']
# cell = sheet.cell(row=19, column=1, )
# cell.value = '123455'

for row in range(1,sheet.max_row):
    for column in range(1, sheet.max_column):
        print(sheet.cell(row=row, column=column).value)
        if row == 19 and column == 1 :
            oldCell = sheet.cell(row=row, column=column)
            oldCell.value = '1234'
            # excel.save('1234.xlsx')
            print('------------------------------------------')
            print(oldCell.value)
rows = [
    ['Number', 'data1', 'data2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10],
]
workBook = Workbook()
ws = workBook.active
ws['a3'] = 'lucifer'
workBook.save('new.xlsx')

